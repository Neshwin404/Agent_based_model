# -*- coding: utf-8 -*-
"""
Created on Sun Nov  6 19:41:13 2022

@author: neshw
"""

import xlrd, xlwt
from xlutils.copy import copy as xl_copy
from openpyxl import load_workbook
# open existing workbook
FilePath = "/Users/farukh/Python ML IVY-May-2020/CarPricesData.xlsx"
 
# Generating workbook
ExcelWorkbook = load_workbook('out_put.xlsx')
# rb = xlrd.open_workbook('out_put.xlsx', formatting_info=True)
# make a copy of it
# wb = xl_copy(rb)
# add sheet to workbook with existing sheets
ExcelWorkbook.create_sheet('sid1')
ExcelWorkbook.save('template.xlsx')
df1.to_excel(ExcelWorkbook, sheet_name = 'sid1')
import pandas as pd
import numpy as np

# path = r"C:\Users\fedel\Desktop\excelData\PhD_data.xlsx"

x1 = np.random.randn(100, 2)
df1 = pd.DataFrame(x1)

x2 = np.random.randn(100, 2)
df2 = pd.DataFrame(x2)

writer = pd.ExcelWriter('out_put.xlsx', engine = 'xlsxwriter')
df1.to_excel(writer, sheet_name = 'x1')
df2.to_excel(writer, sheet_name = 'x2')
writer.save()
writer.close()

book = load_workbook('out_put.xlsx')
writer = pd.ExcelWriter('out_put.xlsx', engine = 'openpyxl')
writer.book = book

x3 = np.random.randn(100, 2)
df3 = pd.DataFrame(x3)

x4 = np.random.randn(100, 2)
df4 = pd.DataFrame(x4)

df3.to_excel(writer, sheet_name = 'x3')
df4.to_excel(writer, sheet_name = 'x4')
writer.close()
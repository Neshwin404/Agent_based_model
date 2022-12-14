"""
Sugarscape Constant Growback Model
================================

Replication of the model found in Netlogo:
Li, J. and Wilensky, U. (2009). NetLogo Sugarscape 2 Constant Growback model.
http://ccl.northwestern.edu/netlogo/models/Sugarscape2ConstantGrowback.
Center for Connected Learning and Computer-Based Modeling,
Northwestern University, Evanston, IL.
"""

import mesa
import numpy as np 
from .agents import SsAgent, Sugar
def gini(arr):
    count = arr.size
    coefficient = 2 / count
    indexes = np.arange(1, count + 1)
    weighted_sum = (indexes * arr).sum()
    total = arr.sum()
    constant = (count + 1) / count
    return coefficient * weighted_sum / total - constant
def get_num_rich_agents(model):
    """return number of rich agents"""
    # print(model.schedule.agents)
    rich_agents = [a.sugar for a in model.schedule.agents  if type(a) is SsAgent ]
    # print(rich_agents)
    rich_agents=np.array(rich_agents)
    # print(rich_agents)
    ap=gini(rich_agents)
    # print(ap)
    return ap

class SugarscapeCg(mesa.Model):
    """
    Sugarscape 2 Constant Growback
    """
    import pandas as pd
    # import xlrd, xlwt
    # from xlutils.copy import copy as xl_copy
    from openpyxl import load_workbook
    verbose = True  # Print-monitoring
    main_data=pd.read_csv("sugarscape_cg/data.csv",index_col=0)
    
    def __init__(self, width=50, height=50, initial_population=main_data.loc[main_data.index[0]].population):
        import pandas as pd
        # import xlrd, xlwt
        # from xlutils.copy import copy as xl_copy
        from openpyxl import load_workbook
        self.rank_dist=[6,3,1]
        self.main_data=pd.read_csv("sugarscape_cg/data.csv",index_col=0)

        """
        Create a new Constant Growback model with the given parameters.

        Args:
            initial_population: Number of population to start with
        """

        # Set parameters
        self.width = width
        self.height = height
        self.initial_population = initial_population

        self.schedule = mesa.time.RandomActivationByType(self)
        self.grid = mesa.space.MultiGrid(self.width, self.height, torus=False)
        # self.datacollector = mesa.DataCollector(
        #     {"SsAgent": lambda m: m.schedule.get_type_count(SsAgent)}
        # )
        self.datacollector = mesa.DataCollector(model_reporters=  {"SsAgent": lambda m: m.schedule.get_type_count(SsAgent),"Rich": get_num_rich_agents},agent_reporters={"Wealth": "sugar","category": "category","vision":"vision"})
        # Create sugar
        import numpy as np
        import pandas as pd
        sugar_distribution = np.genfromtxt("sugarscape_cg/sugar-map.txt")
        self.sdist=sugar_distribution
        # sugar_distribution1=pd.read_csv("sugarscape_cg/sugar_distribution.csv",index_col=0)
        # sugar_distribution1.to_numpy()
        agent_id = 0
        for _, x, y in self.grid.coord_iter():
            max_sugar = sugar_distribution[x, y]
            sugar = Sugar(agent_id, (x, y), self, max_sugar)
            agent_id += 1
            self.grid.place_agent(sugar, (x, y))
            self.schedule.add(sugar)

        # Create agent:
        for i in range(self.initial_population):
            x = self.random.randrange(self.width)
            y = self.random.randrange(self.height)
            sugar = self.random.randrange(6, 25)
            metabolism = self.random.randrange(2, 4)
            vision = self.random.randrange(1, 6)
            tax=0
            ssa = SsAgent(agent_id, (x, y), self, False, sugar, metabolism, vision,tax)
            agent_id += 1
            self.grid.place_agent(ssa, (x, y))
            self.schedule.add(ssa)

        # self.running = True
        self.run_model()
        all_data = self.datacollector.get_agent_vars_dataframe()
        all_data1=self.datacollector.get_model_vars_dataframe()
        all_data=all_data.loc[all_data.index[all_data.Wealth>0]]
        all_data.to_csv("sugar-mapsaaaaass.csv")
        all_data1.to_csv("sugar-mapsaaaaass1.csv")
        self.datacollector.collect(self)
        self.datacollector.collect(self)
        # print(root)

    def step(self):
        self.schedule.step()
        # collect data
        self.datacollector.collect(self)
        if self.verbose:
            print([self.schedule.time, self.schedule.get_type_count(SsAgent)])
    def l_plot(self):
        print("dd")


    def run_model(self, step_count=5):

        if self.verbose:
            print(
                "Initial number Sugarscape Agent: ",
                self.schedule.get_type_count(SsAgent),
            )

        for i in range(step_count):
            self.step()
            self.l_plot()
            
            # global temp_data
            temp_data=self.datacollector.get_agent_vars_dataframe()
            # print(temp_data.columns)
            temp_data=temp_data.loc[temp_data.index[temp_data.category=='ant']]
            if i==(step_count-1):
                temp_data.to_csv("temp_data.csv")
                temp_data=temp_data.reset_index(level=[1])
                print(temp_data.index)
                # temp_data=temp_data.reset_index(level=[0,1])
                import pandas as pd
                # import xlrd, xlwt
                # from xlutils.copy import copy as xl_copy
                from openpyxl import load_workbook
                book = load_workbook('out_put.xlsx')
                writer = pd.ExcelWriter('out_put.xlsx', engine = 'openpyxl')
                writer.book = book
                
                main_data=pd.read_csv("sugarscape_cg/data.csv",index_col=0)
                temp_data.to_excel(writer, sheet_name = main_data.index[0])
                # df4.to_excel(writer, sheet_name = 'x4')
                writer.close()
                
            # temp_data=temp_data.loc[temp_data.index[temp_data.Wealth>0]]
            # temp_data=temp_data[temp_data.Step==i]
            # temp_data=temp_data.sort_values(by=['Wealth'])

            # slab=len(temp_data)/3
            # last_additional=len(temp_data)%3
            # for k in range(3):
            #     print(k)
            #     if k==2:
            #         print("dddddddd")
            #         print(temp_data.loc[3*k:3*(k+1)-1+last_additional])
            #         idx=temp_data[3*k:3*(k+1)-1+last_additional].index
            #         print(idx)
            #         for idn in idx:
            #             temp_data.at[idn,'vision']=6
            #     else:    
            #         print(temp_data.loc[3*k:3*(k+1)-1+last_additional])
            #         idx=temp_data[3*k:3*(k+1)-1+last_additional].index
            #         for idn in idx:
            #             temp_data.at[idn,'vision']=1
                

            # last_additional=len(temp_data)%13
            
            
            
            # temp_data['vision']=0
            # temp_data.loc[(temp_data.sort_values(by=['Wealth'])).index[0:2]]
            # temp_data.to_csv("temp_data.csv")
            # temp_data        

        if self.verbose:
            # print("")
            print(
                "Final number Sugarscape Agent: ",
                self.schedule.get_type_count(SsAgent),
            )
